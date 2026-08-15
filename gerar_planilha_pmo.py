import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date, timedelta
import random

# ── Paleta ──────────────────────────────────────────────────────────
AZUL_ESC  = "002B5C"
AZUL_MED  = "0056A2"
VERDE     = "009A44"
AMARELO   = "FFC200"
VERMELHO  = "C0392B"
CINZA_CLR = "F2F2F2"
CINZA_MED = "CCCCCC"
BRANCO    = "FFFFFF"
PRETO     = "222222"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(color=PRETO, bold=False, size=11, italic=False):
    return Font(name="Calibri", color=color, bold=bold, size=size, italic=italic)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color="002B5C")
    return Border(left=s, right=s, top=s, bottom=s)

def col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def row_height(ws, row, height):
    ws.row_dimensions[row].height = height

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════
# ABA 1 — PROJETOS
# ════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Projetos"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A4"

# Cabeçalho principal
ws1.merge_cells("A1:S1")
ws1["A1"] = "CONTROLE DE PROJETOS — PMO"
ws1["A1"].fill      = fill(AZUL_ESC)
ws1["A1"].font      = font(BRANCO, bold=True, size=16)
ws1["A1"].alignment = center()
row_height(ws1, 1, 36)

ws1.merge_cells("A2:S2")
ws1["A2"] = f"Responsável: Andressa Marquito  |  Atualização: semanal  |  Gerado em: {date.today().strftime('%d/%m/%Y')}"
ws1["A2"].fill      = fill(AZUL_MED)
ws1["A2"].font      = font(BRANCO, size=10, italic=True)
ws1["A2"].alignment = center()
row_height(ws1, 2, 20)

# Cabeçalhos das colunas
headers = [
    "ID", "Nome do Projeto", "Área Demandante", "PMO Responsável",
    "Envolvidos", "Status", "Prioridade",
    "Início Previsto", "Fim Previsto", "Fim Real / Forecast",
    "Orçamento Aprovado (R$)", "Consumido (R$)", "Forecast (R$)",
    "% Consumido", "Qtd. Replanejamentos", "Motivo Último Replanej.",
    "Fase Atual", "Observações", "Última Atualização"
]

for col, h in enumerate(headers, 1):
    c = ws1.cell(row=3, column=col, value=h)
    c.fill      = fill(AZUL_ESC)
    c.font      = font(BRANCO, bold=True, size=10)
    c.alignment = center()
    c.border    = border_thin()
row_height(ws1, 3, 40)

# Larguras
widths = [8, 28, 22, 20, 30, 14, 12, 14, 14, 18, 20, 18, 18, 12, 14, 28, 16, 30, 18]
for i, w in enumerate(widths, 1):
    col_width(ws1, i, w)

# Validações
dv_status = DataValidation(type="list",
    formula1='"🟢 No Prazo,🟡 Atenção,🔴 Crítico,⚫ Encerrado,🔵 Não Iniciado"',
    allow_blank=True)
dv_prior = DataValidation(type="list",
    formula1='"Alta,Média,Baixa"', allow_blank=True)
dv_fase = DataValidation(type="list",
    formula1='"Iniciação,Planejamento,Execução,Monitoramento,Encerramento"',
    allow_blank=True)
ws1.add_data_validation(dv_status)
ws1.add_data_validation(dv_prior)
ws1.add_data_validation(dv_fase)

# Dados de exemplo
projetos = [
    ["PRJ-001","Implantação ERP Financeiro","Financeiro","Ana Paula","TI, Financeiro, Controladoria","🟢 No Prazo","Alta","01/02/2025","31/08/2025","31/08/2025",850000,420000,860000,0,0,"—","Execução","Dentro do prazo e orçamento"],
    ["PRJ-002","Portal do Cliente","Comercial","Bruno Silva","TI, Comercial, Marketing","🟡 Atenção","Alta","15/01/2025","30/06/2025","15/07/2025",320000,280000,345000,0,1,"Escopo ampliado","Execução","Risco de estouro de orçamento"],
    ["PRJ-003","Automação de Relatórios","PMO","Carla Matos","PMO, TI","🟢 No Prazo","Média","01/03/2025","30/09/2025","30/09/2025",180000,75000,180000,0,0,"—","Planejamento","Sem impedimentos"],
    ["PRJ-004","Migração Cloud","TI","Daniel Costa","TI, Infraestrutura","🔴 Crítico","Alta","01/11/2024","28/02/2025","30/06/2025",600000,590000,720000,0,2,"Problemas técnicos e atraso fornecedor","Execução","Projeto em situação crítica — escalado"],
    ["PRJ-005","Programa de Agilidade","RH / PMO","Ana Paula","PMO, RH, Líderes","🟡 Atenção","Média","01/02/2025","31/12/2025","31/12/2025",250000,85000,265000,0,0,"—","Execução","Engajamento abaixo do esperado"],
    ["PRJ-006","Novo Produto Vida","Produtos","Bruno Silva","Produtos, Atuária, Compliance","🟢 No Prazo","Alta","01/04/2025","30/11/2025","30/11/2025",430000,120000,430000,0,0,"—","Planejamento","Aprovação regulatória em andamento"],
    ["PRJ-007","Integração Parceiro Externo","TI","Carla Matos","TI, Jurídico, Parceiro","🟡 Atenção","Alta","15/02/2025","31/07/2025","15/08/2025",290000,195000,310000,0,1,"Atraso na entrega do parceiro","Execução","Comunicação formal enviada ao parceiro"],
    ["PRJ-008","Dashboard Executivo","PMO","Daniel Costa","PMO, TI, Diretoria","🟢 No Prazo","Média","01/03/2025","31/05/2025","31/05/2025",95000,60000,95000,0,0,"—","Execução","Entrega prevista no prazo"],
    ["PRJ-009","Reestruturação de Processos","Operações","Ana Paula","Operações, PMO, RH","🔵 Não Iniciado","Baixa","01/06/2025","31/12/2025","31/12/2025",200000,0,200000,0,0,"—","Iniciação","Aguardando kick-off"],
    ["PRJ-010","Modernização de Infraestrutura","TI","Bruno Silva","TI, Infraestrutura","🟢 No Prazo","Média","01/01/2025","30/09/2025","30/09/2025",510000,280000,510000,0,0,"—","Execução","Conforme planejado"],
    ["PRJ-011","App Mobile Clientes","Comercial","Carla Matos","TI, Comercial, UX","🟡 Atenção","Alta","01/03/2025","31/10/2025","30/11/2025",380000,140000,395000,0,1,"Mudança de escopo UX","Execução","Nova previsão aprovada pela diretoria"],
    ["PRJ-012","Compliance LGPD","Jurídico","Daniel Costa","Jurídico, TI, RH","🟢 No Prazo","Alta","01/11/2024","30/04/2025","30/04/2025",160000,155000,162000,0,0,"—","Monitoramento","Entrega final em revisão"],
    ["PRJ-013","Treinamento Liderança","RH","Ana Paula","RH, Líderes","🔵 Não Iniciado","Baixa","01/07/2025","31/12/2025","31/12/2025",80000,0,80000,0,0,"—","Iniciação","Cronograma em elaboração"],
    ["PRJ-014","Renovação Contratos TI","Jurídico / TI","Bruno Silva","Jurídico, TI, Suprimentos","🟢 No Prazo","Média","15/01/2025","15/06/2025","15/06/2025",50000,35000,50000,0,0,"—","Execução","Contratos em fase de assinatura"],
    ["PRJ-015","Plano Estratégico 2026","Diretoria","Carla Matos","Diretoria, PMO, Todas as áreas","🟡 Atenção","Alta","01/04/2025","31/10/2025","15/11/2025",120000,35000,125000,0,1,"Revisão de escopo pela diretoria","Planejamento","Reunião de alinhamento agendada"],
]

status_color = {
    "🟢 No Prazo":    "E8F5E9",
    "🟡 Atenção":     "FFFDE7",
    "🔴 Crítico":     "FFEBEE",
    "⚫ Encerrado":   "EEEEEE",
    "🔵 Não Iniciado":"E3F2FD",
}

for r, p in enumerate(projetos, 4):
    row_color = status_color.get(p[5], BRANCO)
    for c, val in enumerate(p, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.fill      = fill(row_color if r % 2 == 0 else BRANCO)
        cell.font      = font(size=10)
        cell.alignment = center() if c not in [2,4,5,16,18] else left()
        cell.border    = border_thin()

    # Calcula % consumido
    orc = p[10]; cons = p[11]
    pct = round(cons / orc * 100, 1) if orc else 0
    ws1.cell(row=r, column=14).value = pct / 100
    ws1.cell(row=r, column=14).number_format = "0.0%"

    # Formata moeda
    for col_idx in [11, 12, 13]:
        ws1.cell(row=r, column=col_idx).number_format = 'R$ #,##0'

    # Data de atualização
    ws1.cell(row=r, column=19).value = date.today().strftime("%d/%m/%Y")

    dv_status.add(ws1.cell(row=r, column=6))
    dv_prior.add(ws1.cell(row=r, column=7))
    dv_fase.add(ws1.cell(row=r, column=17))

row_height(ws1, *[r for r in range(4, 20)]) if False else None
for r in range(4, 20):
    row_height(ws1, r, 28)

# ════════════════════════════════════════════════════════════════════
# ABA 2 — DASHBOARD
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Dashboard")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:L1")
ws2["A1"] = "DASHBOARD EXECUTIVO — PMO"
ws2["A1"].fill      = fill(AZUL_ESC)
ws2["A1"].font      = font(BRANCO, bold=True, size=16)
ws2["A1"].alignment = center()
row_height(ws2, 1, 36)

ws2.merge_cells("A2:L2")
ws2["A2"] = f"Andressa Marquito  |  {date.today().strftime('%d/%m/%Y')}  |  Atualização: semanal"
ws2["A2"].fill      = fill(AZUL_MED)
ws2["A2"].font      = font(BRANCO, size=10, italic=True)
ws2["A2"].alignment = center()
row_height(ws2, 2, 20)

# KPI Cards
kpis_headers = ["KPI", "Valor", "Observação"]
kpis = [
    ["Total de Projetos", 15, ""],
    ["🟢 No Prazo", 6, "40%"],
    ["🟡 Atenção", 5, "33%"],
    ["🔴 Crítico", 1, "7%"],
    ["🔵 Não Iniciados", 2, "13%"],
    ["⚫ Encerrados", 1, "7%"],
    ["Orçamento Total Aprovado", "R$ 4.513.000", ""],
    ["Total Consumido", "R$ 2.465.000", "55% do total"],
    ["Forecast Total", "R$ 4.727.000", "+R$ 214k vs. aprovado"],
    ["Projetos com Replanej.", 5, "33% do portfólio"],
    ["PMOs ativos", 4, "Ana, Bruno, Carla, Daniel"],
]

ws2.merge_cells("A4:C4")
ws2["A4"] = "INDICADORES DO PORTFÓLIO"
ws2["A4"].fill      = fill(AZUL_ESC)
ws2["A4"].font      = font(BRANCO, bold=True, size=12)
ws2["A4"].alignment = center()
row_height(ws2, 4, 28)

for i, (kpi, val, obs) in enumerate(kpis, 5):
    bg = CINZA_CLR if i % 2 == 0 else BRANCO
    for c, v in enumerate([kpi, val, obs], 1):
        cell = ws2.cell(row=i, column=c, value=v)
        cell.fill      = fill(bg)
        cell.font      = font(size=11, bold=(c == 1))
        cell.alignment = center() if c > 1 else left()
        cell.border    = border_thin()
    row_height(ws2, i, 24)

ws2.column_dimensions["A"].width = 32
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 28

# Tabela Status por área
ws2.merge_cells("E4:H4")
ws2["E4"] = "STATUS POR ÁREA DEMANDANTE"
ws2["E4"].fill      = fill(AZUL_ESC)
ws2["E4"].font      = font(BRANCO, bold=True, size=12)
ws2["E4"].alignment = center()

area_headers = ["Área", "Total", "🟢", "🟡", "🔴"]
for c, h in enumerate(area_headers, 5):
    cell = ws2.cell(row=5, column=c, value=h)
    cell.fill      = fill(AZUL_MED)
    cell.font      = font(BRANCO, bold=True, size=10)
    cell.alignment = center()
    cell.border    = border_thin()

areas_data = [
    ["TI", 4, 2, 1, 1],
    ["Comercial", 2, 1, 1, 0],
    ["Financeiro", 1, 1, 0, 0],
    ["PMO", 2, 2, 0, 0],
    ["Produtos", 1, 1, 0, 0],
    ["RH", 2, 0, 1, 0],
    ["Jurídico", 2, 2, 0, 0],
    ["Diretoria", 1, 0, 1, 0],
]
for r, row in enumerate(areas_data, 6):
    bg = CINZA_CLR if r % 2 == 0 else BRANCO
    for c, v in enumerate(row, 5):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.fill      = fill(bg)
        cell.font      = font(size=10)
        cell.alignment = center()
        cell.border    = border_thin()
    row_height(ws2, r, 22)

for c in range(5, 10):
    ws2.column_dimensions[get_column_letter(c)].width = 14

# Tabela PMOs
ws2.merge_cells("J4:L4")
ws2["J4"] = "PROJETOS POR PMO"
ws2["J4"].fill      = fill(AZUL_ESC)
ws2["J4"].font      = font(BRANCO, bold=True, size=12)
ws2["J4"].alignment = center()

pmo_headers = ["PMO", "Projetos", "Críticos"]
for c, h in enumerate(pmo_headers, 10):
    cell = ws2.cell(row=5, column=c, value=h)
    cell.fill      = fill(AZUL_MED)
    cell.font      = font(BRANCO, bold=True, size=10)
    cell.alignment = center()
    cell.border    = border_thin()

pmos = [
    ["Ana Paula", 4, 0],
    ["Bruno Silva", 4, 0],
    ["Carla Matos", 4, 1],
    ["Daniel Costa", 3, 0],
]
for r, row in enumerate(pmos, 6):
    bg = CINZA_CLR if r % 2 == 0 else BRANCO
    for c, v in enumerate(row, 10):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.fill      = fill(bg)
        cell.font      = font(size=10,
                              color=VERMELHO if (c == 12 and v > 0) else PRETO,
                              bold=(c == 12 and v > 0))
        cell.alignment = center()
        cell.border    = border_thin()
    row_height(ws2, r, 22)

for c in range(10, 13):
    ws2.column_dimensions[get_column_letter(c)].width = 18

# ════════════════════════════════════════════════════════════════════
# ABA 3 — DOCUMENTOS
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Documentos")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:J1")
ws3["A1"] = "CONTROLE DE DOCUMENTOS POR PROJETO"
ws3["A1"].fill      = fill(AZUL_ESC)
ws3["A1"].font      = font(BRANCO, bold=True, size=14)
ws3["A1"].alignment = center()
row_height(ws3, 1, 32)

ws3.merge_cells("A2:J2")
ws3["A2"] = "Registre aqui todos os documentos vinculados a cada projeto"
ws3["A2"].fill      = fill(AZUL_MED)
ws3["A2"].font      = font(BRANCO, size=10, italic=True)
ws3["A2"].alignment = center()
row_height(ws3, 2, 18)

doc_headers = [
    "ID Projeto", "Nome do Projeto", "Tipo de Documento",
    "Nome do Arquivo / Documento", "Versão", "Status",
    "Data de Emissão", "Responsável", "Aprovado por", "Observações"
]
for c, h in enumerate(doc_headers, 1):
    cell = ws3.cell(row=3, column=c, value=h)
    cell.fill      = fill(AZUL_ESC)
    cell.font      = font(BRANCO, bold=True, size=10)
    cell.alignment = center()
    cell.border    = border_thin()
row_height(ws3, 3, 36)

dv_doc_status = DataValidation(type="list",
    formula1='"Rascunho,Em Revisão,Aprovado,Obsoleto"', allow_blank=True)
dv_doc_tipo = DataValidation(type="list",
    formula1='"Termo de Abertura,Plano do Projeto,Ata de Reunião,Cronograma,Orçamento,Relatório de Status,Contrato,Proposta,Outro"',
    allow_blank=True)
ws3.add_data_validation(dv_doc_status)
ws3.add_data_validation(dv_doc_tipo)

docs = [
    ["PRJ-001","Implantação ERP Financeiro","Termo de Abertura","TAP_ERP_Financeiro_v1.pdf","1.0","Aprovado","10/01/2025","Ana Paula","Diretoria",""],
    ["PRJ-001","Implantação ERP Financeiro","Cronograma","Cronograma_ERP_v2.xlsx","2.0","Aprovado","15/01/2025","Ana Paula","PMO","Replanejado em mar/25"],
    ["PRJ-002","Portal do Cliente","Termo de Abertura","TAP_Portal_v1.pdf","1.0","Aprovado","20/01/2025","Bruno Silva","Diretoria",""],
    ["PRJ-002","Portal do Cliente","Plano do Projeto","PP_Portal_v1.docx","1.0","Em Revisão","25/01/2025","Bruno Silva","Ana Paula",""],
    ["PRJ-004","Migração Cloud","Termo de Abertura","TAP_Cloud_v1.pdf","1.0","Aprovado","05/10/2024","Daniel Costa","CIO",""],
    ["PRJ-004","Migração Cloud","Relatório de Status","Status_Cloud_jun25.pdf","6.0","Aprovado","01/06/2025","Daniel Costa","PMO","Relatório de escalação"],
    ["PRJ-012","Compliance LGPD","Plano do Projeto","PP_LGPD_v2.docx","2.0","Aprovado","10/11/2024","Daniel Costa","Jurídico",""],
    ["PRJ-012","Compliance LGPD","Contrato","Contrato_DPO_v1.pdf","1.0","Aprovado","15/11/2024","Daniel Costa","Jurídico",""],
]

doc_status_color = {
    "Aprovado":    "E8F5E9",
    "Em Revisão":  "FFFDE7",
    "Rascunho":    "F3E5F5",
    "Obsoleto":    "EEEEEE",
}
for r, d in enumerate(docs, 4):
    bg = doc_status_color.get(d[5], BRANCO)
    for c, v in enumerate(d, 1):
        cell = ws3.cell(row=r, column=c, value=v)
        cell.fill      = fill(bg)
        cell.font      = font(size=10)
        cell.alignment = center() if c not in [2, 4, 10] else left()
        cell.border    = border_thin()
    dv_doc_status.add(ws3.cell(row=r, column=6))
    dv_doc_tipo.add(ws3.cell(row=r, column=3))
    row_height(ws3, r, 24)

doc_widths = [10, 28, 22, 32, 8, 14, 16, 18, 18, 28]
for i, w in enumerate(doc_widths, 1):
    col_width(ws3, i, w)

# ════════════════════════════════════════════════════════════════════
# ABA 4 — REPLANEJAMENTOS
# ════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Replanejamentos")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:J1")
ws4["A1"] = "HISTÓRICO DE REPLANEJAMENTOS"
ws4["A1"].fill      = fill(AZUL_ESC)
ws4["A1"].font      = font(BRANCO, bold=True, size=14)
ws4["A1"].alignment = center()
row_height(ws4, 1, 32)

ws4.merge_cells("A2:J2")
ws4["A2"] = "Registre cada replanejamento — prazo, orçamento ou escopo"
ws4["A2"].fill      = fill(AZUL_MED)
ws4["A2"].font      = font(BRANCO, size=10, italic=True)
ws4["A2"].alignment = center()
row_height(ws4, 2, 18)

rep_headers = [
    "ID Projeto", "Nome do Projeto", "Nº Replanej.", "Data",
    "Tipo", "Prazo Anterior", "Novo Prazo",
    "Orçamento Anterior (R$)", "Novo Orçamento (R$)", "Motivo / Justificativa"
]
for c, h in enumerate(rep_headers, 1):
    cell = ws4.cell(row=3, column=c, value=h)
    cell.fill      = fill(AZUL_ESC)
    cell.font      = font(BRANCO, bold=True, size=10)
    cell.alignment = center()
    cell.border    = border_thin()
row_height(ws4, 3, 36)

reps = [
    ["PRJ-002","Portal do Cliente",1,"15/03/2025","Escopo + Prazo","30/06/2025","15/07/2025",320000,345000,"Ampliação de escopo solicitada pela diretoria comercial"],
    ["PRJ-004","Migração Cloud",1,"20/01/2025","Prazo","28/02/2025","31/05/2025",600000,600000,"Problemas técnicos na migração do ambiente de homologação"],
    ["PRJ-004","Migração Cloud",2,"10/04/2025","Prazo + Orçamento","31/05/2025","30/06/2025",600000,720000,"Atraso do fornecedor + horas extras necessárias"],
    ["PRJ-007","Integração Parceiro Externo",1,"01/04/2025","Prazo","31/07/2025","15/08/2025",290000,310000,"Parceiro externo não entregou API conforme contrato"],
    ["PRJ-011","App Mobile Clientes",1,"10/05/2025","Escopo + Prazo","31/10/2025","30/11/2025",380000,395000,"Mudança de escopo de UX aprovada pelo comitê"],
    ["PRJ-015","Plano Estratégico 2026",1,"20/05/2025","Escopo + Prazo","31/10/2025","15/11/2025",120000,125000,"Revisão de escopo solicitada pela presidência"],
]

for r, rep in enumerate(reps, 4):
    bg = CINZA_CLR if r % 2 == 0 else BRANCO
    for c, v in enumerate(rep, 1):
        cell = ws4.cell(row=r, column=c, value=v)
        cell.fill      = fill(bg)
        cell.font      = font(size=10)
        cell.alignment = center() if c != 10 else left()
        cell.border    = border_thin()
        if c in [8, 9]:
            cell.number_format = 'R$ #,##0'
    row_height(ws4, r, 28)

rep_widths = [10, 28, 12, 14, 18, 16, 16, 22, 22, 40]
for i, w in enumerate(rep_widths, 1):
    col_width(ws4, i, w)

# ════════════════════════════════════════════════════════════════════
# ABA 5 — INSTRUÇÕES
# ════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Instruções")
ws5.sheet_view.showGridLines = False

ws5.merge_cells("A1:D1")
ws5["A1"] = "INSTRUÇÕES DE USO — CONTROLE PMO"
ws5["A1"].fill      = fill(AZUL_ESC)
ws5["A1"].font      = font(BRANCO, bold=True, size=14)
ws5["A1"].alignment = center()
row_height(ws5, 1, 32)

instrucoes = [
    ("ABA: Projetos", "Cadastro completo de cada projeto. Preencha todas as colunas. Status, Prioridade e Fase possuem listas suspensas — clique na célula para ver as opções. Orçamento em valores reais (R$). A coluna '% Consumido' calcula automaticamente."),
    ("ABA: Dashboard", "Visão executiva consolidada. Atualize manualmente os KPIs semanalmente ou após mudanças relevantes no portfólio. Use para apresentações à diretoria."),
    ("ABA: Documentos", "Registre todos os documentos de cada projeto: Tipo, Versão, Status e Responsável. Status com lista suspensa: Rascunho → Em Revisão → Aprovado → Obsoleto."),
    ("ABA: Replanejamentos", "Registre CADA replanejamento individualmente — prazo, orçamento ou escopo. Inclua sempre o motivo. O número sequencial por projeto permite rastrear o histórico completo."),
    ("Atualização semanal", "Toda segunda-feira: atualize o status de cada projeto, registre novos documentos e replanejamentos ocorridos na semana. Compartilhe o arquivo atualizado com a equipe via e-mail ou pasta compartilhada."),
    ("Campos obrigatórios", "ID, Nome do Projeto, Área Demandante, PMO Responsável, Status, Orçamento Aprovado e Datas são obrigatórios para todos os projetos."),
    ("Codificação de IDs", "Use o padrão PRJ-001, PRJ-002... para projetos. Para documentos, use o ID do projeto como referência. Para replanejamentos, numere sequencialmente por projeto (1, 2, 3...)."),
    ("Backup", "Salve uma cópia semanal com a data no nome: ex. Controle_PMO_20250815.xlsx. Mantenha as últimas 4 semanas."),
]

for r, (titulo, texto) in enumerate(instrucoes, 3):
    ws5.cell(row=r, column=1, value=titulo).font = font(BRANCO, bold=True, size=11)
    ws5.cell(row=r, column=1).fill = fill(AZUL_MED)
    ws5.cell(row=r, column=1).alignment = left()
    ws5.cell(row=r, column=1).border = border_thin()

    ws5.merge_cells(f"B{r}:D{r}")
    c = ws5.cell(row=r, column=2, value=texto)
    c.font      = font(size=10)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.fill      = fill(CINZA_CLR if r % 2 == 0 else BRANCO)
    c.border    = border_thin()
    row_height(ws5, r, 52)

ws5.column_dimensions["A"].width = 26
ws5.column_dimensions["B"].width = 80

# ── Ordem das abas ────────────────────────────────────────────────
wb._sheets = [ws1, ws2, ws3, ws4, ws5]

# ── Salvar ────────────────────────────────────────────────────────
output = r"c:\Users\dessa\OneDrive\Área de Trabalho\Controle_PMO_Andressa.xlsx"
wb.save(output)
print(f"Arquivo salvo: {output}")
